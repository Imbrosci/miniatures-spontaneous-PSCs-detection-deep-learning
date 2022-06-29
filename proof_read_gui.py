# -*- coding: utf-8 -*-
"""
Created on Tue Feb 15 09:57:32 2022.

@author: imbroscb
"""

import warnings
warnings.filterwarnings("ignore")
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
import pandas as pd
from sklearn.metrics.pairwise import euclidean_distances
from pscs_parameters_calculator import ParametersCalculator


class MiniSpontProof:
    """Give the possibility to check and revise the results."""

    def __init__(self, root, results_path, sheet_name, time, signal_lp,
                 original_values, fs):

        self.root = root
        self.results_path = results_path
        self.sheet_name = sheet_name
        self.time = time
        self.signal_lp = signal_lp
        self.original_values = original_values
        self.fs = fs
        self.init_window()

    def init_window(self):
        """Plot signal_lp and detected m/s PSCs."""
        self.fig = plt.figure(figsize=(16, 8))
        self.ax = self.fig.add_subplot(1, 1, 1)
        self.ax.plot(self.time, self.signal_lp, c='black')
        if len(self.original_values) > 0:
            self.ax.scatter(self.original_values[:, 0],
                            self.original_values[:, 1], c='orange', s=70)
        self.ax.title.set_text('Press "u" to update the results.xlsx file.')

        # the following lines are necessary for the manual correction
        self.updated_values = self.original_values
        self.false_positives = []
        plt.connect('key_press_event', self.create_new_dataset)
        plt.show()

    def create_new_dataset(self, event):
        """
        Add points, remove points and update the results.xlsx file.

        The keyboard button 'a' is used to add false negatives.
        The keyboard button 'd' is used to remove false positives.
        The keyboard button 'u' is used to update the results.xlsx file.
        """
        # add a new event in response to the keyboard button 'a'
        if event.key == 'a':
            new_xy_event = [event.xdata, event.ydata]

            # define the current events
            xy_events_tp = self.updated_values

            # update new values: current events + new event
            if len(self.updated_values) > 0:
                self.updated_values = np.insert(xy_events_tp, 0, new_xy_event,
                                                axis=0)
            else:
                self.updated_values = np.array(new_xy_event).reshape(1, -1)

            # plot the new event
            self.ax.scatter(event.xdata, event.ydata, c='blue', s=70)
            self.fig.canvas.draw()

        # delete an event in response to the keyboard button 'd',
        # if a data point is close enough to the clicking position
        if event.key == 'd':

            xy_click = np.array([event.xdata, event.ydata])
            xy_events = self.updated_values

            # look for event detection (data point) closest to the click
            closest_point = xy_events[0]
            min_euc = 999999
            idx_to_delete = 0
            for idx in range(len(xy_events)):
                euc_distance = euclidean_distances(
                    xy_events[idx].reshape(1, -1), xy_click.reshape(1, -1))
                if euc_distance < min_euc:
                    min_euc = euc_distance
                    closest_point = xy_events[idx]
                    idx_to_delete = idx

            # if the closest data point is close enough, it will be deleted
            if min_euc < 50:
                self.updated_values = np.delete(xy_events, idx_to_delete,
                                                axis=0)
                # plot the the deleted data point in red
                self.ax.scatter(closest_point[0], closest_point[1], c='red',
                                s=70)
                self.fig.canvas.draw()

                # define the current false positives data point
                xy_events_fp = self.false_positives

                # update false positives
                if len(self.false_positives) > 0:
                    self.false_positives = np.insert(xy_events_fp, 0,
                                                     closest_point, axis=0)
                else:
                    self.false_positives = np.array(
                        closest_point).reshape(1, -1)

        if event.key == 'u':

            # get the updated x, y values
            x = self.updated_values[:, 0].astype('float64')
            y = self.updated_values[:, 1].astype('float64')

            # sort the updated x, y values (for ascending x)
            x, y = zip(*sorted(zip(x, y)))
            x = np.array(x)
            y = np.array(y)

            # get the datapoint from the ms
            x_dp = (x * (self.fs / 1000)).astype('int')

            # calculate the interevent interval
            x_pd = pd.Series(x)
            interevent_interval = x_pd.shift(-1) - x_pd
            average_interevent_interval = np.nanmean(interevent_interval)

            # instantiate parameters
            parameters = ParametersCalculator(self.signal_lp, x_dp, y, self.fs)

            # calculate the amplitude
            amplitude = parameters.amplitude_stdev()[0]
            average_amplitude = np.nanmean(amplitude)

            # calculate the mean rise (10-90%) and the mean decay time (90-10%)
            mean_rise_time = parameters.rise_time()[0]
            mean_decay_time = parameters.decay_time()[0]

            # calculate the stdev of the baseline signal
            stdev = parameters.amplitude_stdev()[1]
            average_stdev = np.nanmean(stdev)

            # load the excel file
            wb = load_workbook(self.root + self.results_path)

            # get header from the sheet with original values
            ws_to_remove = wb[self.sheet_name]
            a1 = ws_to_remove.cell(1, 1).value
            b1 = ws_to_remove.cell(1, 2).value
            c1 = ws_to_remove.cell(1, 3).value
            d1 = ws_to_remove.cell(1, 4).value

            # remove the sheet with original values
            wb.remove(ws_to_remove)

            # create a sheet with the same title
            ws_revised = wb.create_sheet()
            ws_revised.title = self.sheet_name

            # fill the created sheet with the revised values
            ws_revised.cell(1, 1).value = a1
            ws_revised.cell(1, 2).value = b1
            ws_revised.cell(1, 3).value = c1
            ws_revised.cell(1, 4).value = d1

            for row in range(len(x)):
                ws_revised.cell(row + 2, 1).value = x[row]
                ws_revised.cell(row + 2, 2).value = y[row]
                ws_revised.cell(row + 2, 3).value = interevent_interval[row]
                ws_revised.cell(row + 2, 4).value = amplitude[row]

            # modify the Summary results sheet
            ws_summary = wb['Summary results']

            for col in range(ws_summary.max_column):
                if ws_summary.cell(1, col + 1).value == 'Recording filename':
                    col_rec_file = col + 1
                if ws_summary.cell(1, col + 1).value == 'Channel':
                    col_channel = col + 1
                if ws_summary.cell(1, col + 1).value == 'Average interevent interval (ms)':
                    col_inter_ev_inter = col + 1
                if ws_summary.cell(1, col + 1).value == 'Average amplitude (pA)':
                    col_amplitude = col + 1
                if ws_summary.cell(1,col + 1).value == 'Average 10-90% rise time (ms)':
                    col_rise_time = col + 1
                if ws_summary.cell(1, col + 1).value == 'Average 90-10% decay time (ms)':
                    col_decay_time = col + 1
                if ws_summary.cell(1, col + 1).value == 'Stdev of the baseline signal (pA)':
                    col_stdev = col + 1
                if ws_summary.cell(1, col + 1).value == 'Manually revised':
                    col_man_rev = col + 1

            for row in range(ws_summary.max_row):
                if ws_summary.cell(row + 1, col_rec_file).value[:-4] == self.sheet_name[:-2]:
                    if ws_summary.cell(row + 1,col_channel).value == int(self.sheet_name.split('_')[-1]):
                        row_to_modify = row + 1

            ws_summary.cell(
                row_to_modify,
                col_inter_ev_inter).value = average_interevent_interval
            ws_summary.cell(
                row_to_modify, col_amplitude).value = average_amplitude
            ws_summary.cell(
                row_to_modify, col_rise_time).value = mean_rise_time
            ws_summary.cell(
                row_to_modify, col_decay_time).value = mean_decay_time
            ws_summary.cell(row_to_modify, col_stdev).value = average_stdev
            ws_summary.cell(row_to_modify, col_man_rev).value = 'yes'

            # save and close wb
            wb.save(self.root + self.results_path)
            print('The results.xlsx file has been updated')
            wb.close()
